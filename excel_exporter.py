"""
Module for exporting QnA pairs to Excel file
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict
import os


class ExcelExporter:
    """Export QnA pairs to Excel format"""
    
    def __init__(self, output_path: str = "Multilingual_QnA.xlsx"):
        """
        Initialize Excel Exporter
        
        Args:
            output_path: Path where the Excel file will be saved
        """
        self.output_path = output_path
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
    
    def add_sheet(self, sheet_name: str, qna_pairs: List[Dict[str, str]]) -> None:
        """
        Add a new sheet with QnA pairs
        
        Args:
            sheet_name: Name of the sheet (e.g., 'English', 'Hindi', 'Marathi')
            qna_pairs: List of QnA dictionaries with 'question' and 'answer' keys
        """
        worksheet = self.workbook.create_sheet(sheet_name)
        
        # Add headers
        headers = ['Questions', 'Answers']
        worksheet.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add QnA pairs
        for qna in qna_pairs:
            worksheet.append([
                qna.get('question', ''),
                qna.get('answer', '')
            ])
        
        # Style cells and set column widths
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        worksheet.column_dimensions['A'].width = 50
        worksheet.column_dimensions['B'].width = 60
        
        for row in worksheet.iter_rows(min_row=2, max_row=len(qna_pairs) + 1, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                # Set row height for better readability
                cell.parent.height = None  # Auto height
        
        # Set header row height
        worksheet.row_dimensions[1].height = 25
    
    def save(self) -> str:
        """
        Save the Excel file
        
        Returns:
            Path to the saved file
        """
        try:
            self.workbook.save(self.output_path)
            return self.output_path
        except Exception as e:
            raise Exception(f"Error saving Excel file: {str(e)}")
    
    def create_multilingual_qna_file(self, 
                                     english_qna: List[Dict[str, str]],
                                     hindi_qna: List[Dict[str, str]],
                                     marathi_qna: List[Dict[str, str]]) -> str:
        """
        Create a complete multilingual QnA Excel file
        
        Args:
            english_qna: List of QnA pairs in English
            hindi_qna: List of QnA pairs in Hindi
            marathi_qna: List of QnA pairs in Marathi
            
        Returns:
            Path to the created file
        """
        self.add_sheet("English", english_qna)
        self.add_sheet("Hindi", hindi_qna)
        self.add_sheet("Marathi", marathi_qna)
        
        return self.save()
